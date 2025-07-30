from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from typing import List
from io import BytesIO


def iniciar_planilha(conteudo):
    """
    Inicia o arquivo Excel contendo os dados.
    :param conteudo: Arquivo da planilha excel, podendo ser um path ou o arquivo em bytes.
    :return: Retorna a planilha.
    """
    try:
        planilha = load_workbook(filename=conteudo, data_only=True)
    except PermissionError:
        print('O arquivo já está aberto, feche o mesmo antes de prosseguir.')
        exit()
    else:
        return planilha


def descobrir_linha_vazia_planilha(conteudo) -> str:
    """
    Descobre o número da ultima linha preenchida na planilha da coluna informada.
    Mesmo se houver linhas vazias entres linhas preenchidas, a última será pega.
    Versão extremamente otimizada para obter a ultima linha.
    :param conteudo: Arquivo da planilha excel, podendo ser um path ou o arquivo em bytes.
    :return: Retorna o número da ultima linha como uma string.
    """
    # 1) Abre em modo normal (não read_only, porque queremos ws._cells)
    planilha = iniciar_planilha(conteudo)
    aba_ativa = planilha.active

    # # 2) Converte letra de coluna em índice numérico
    # coluna_idx = column_index_from_string(coluna.upper())

    # print(list(ws[coluna])[-1].row)
    try:
        # 3) Pega todas as chaves (row, col) que o OpenPyXL carrega em _cells
        # e filtra só as que têm coluna = col_idx
        # coord é uma tupla (row, col)
        # ultima_linha = max([row for (row, col) in ws._cells.keys()], default=1)
        ultima_linha = max([cell.row for cell in aba_ativa._cells.values() if cell.value is not None], default=1)
        # default=1 -> Normalmente é a linha onde eu começo a preencher.

        # print(ws._cells)
        # print(ws._cells.keys())
        # print(ws._cells.values())

        return str(ultima_linha)
    except AttributeError:
        # 3.1) fallback do métod0 oficial (Percorre todas as células visíveis na planilha):
        ultima_linha = 1
        for row in aba_ativa.iter_rows():
            for cell in row:
                if cell.value is not None:
                    ultima_linha = max(ultima_linha, cell.row)
        return str(ultima_linha)
    finally:
        planilha.close()


def is_merged(cell, merged_ranges):
    """ Verifica se a célula faz parte de uma faixa mesclada """
    for merged_range in merged_ranges:
        # cell.coordinate: Retorna a referência da célula (por exemplo, 'A1', 'B2', etc.).
        if cell.coordinate in merged_range:  # Testa se a referência da célula atual está dentro da faixa mesclada.
            return True
    return False


def converter_objeto_para_planilha(data: List[dict]) -> bytes:
    """
    Converte um objeto para uma planilha.
    :param data:
    :return: Retorna a planilha no formato de bytes.
    """
    # Cria um novo workbook
    wb = Workbook()
    ws = wb.active

    # Adiciona o cabeçalho ao array "headers":
    headers = list(dict(data[0]).keys()) if data else []

    # Aplica o estilo negrito da planilha em uma variável:
    bold_font = Font(bold=True)

    # Transformando todos os campos do formato "campoTeste" para "CAMPO TESTE":
    headers = [camel_case_to_upper_case_with_spaces(element) for element in headers]

    # Adicionando o cabeçalho na planilha:
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header.upper())
        cell.font = bold_font  # Define a celula atual da planilha como negrito.

    # Adicionando os valores nas linhas:
    for row in data:
        ws.append(list(dict(row).values()))

    # Salva o workbook em um BytesIO buffer
    buffer = BytesIO()
    wb.save(buffer)

    # Move o ponteiro de volta para o início, permitindo que o conteúdo do buffer seja lido corretamente.
    # Caso eu tentasse usar o .getvalue sem fazer o .seek(0) o valor lido seria vazio, pois o ponteiro estaria
    # no final do conteudo.
    buffer.seek(0)

    # Retorna o arquivo Excel como uma resposta
    return buffer.getvalue()


def camel_case_to_upper_case_with_spaces(s: str):
    """
    Função responsável por transformar "camelCase" para "UPPER CASE WITH SPACES".
    :param s: String que será transformada.
    :return: String transformada.
    """
    import re

    # Insere um espaço antes de cada letra maiúscula
    s_with_spaces = re.sub(r'(?<!^)(?=[A-Z])', ' ', s)
    # Converte para maiúsculas
    return s_with_spaces.upper()
