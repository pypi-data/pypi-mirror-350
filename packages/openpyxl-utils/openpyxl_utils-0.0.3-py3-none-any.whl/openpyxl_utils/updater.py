from openpyxl_utils.utils import iniciar_planilha


def atualizar_dados_intervalo_planilha(conteudo, valores_adicionar: list, intervalo: str) -> None:
    """
    Atualiza os dados presentes no intervalo informado.
    Caso o intervalo esteja vazio, um erro de Index é gerado.  <- FUNCIONALIDADE DESATIVADA
    :param conteudo: Arquivo da planilha excel, podendo ser um path ou o arquivo em bytes.
    :param valores_adicionar: Lista de listas contendo os valores que vão substituir os dados presentes no intervalo.
    :param intervalo: Intervalo que será substituído.
    :return: None
    """
    planilha = iniciar_planilha(conteudo)
    aba_ativa = planilha.active

    try:
        for i, linha in enumerate(aba_ativa[intervalo]):  # -> A2, B2, C2, ...
            if i >= len(valores_adicionar):
                break
            for j, elemento in enumerate(linha):
                # if elemento.value is None:  # -> Gerando o erro de função específica.
                    # raise IndexError("Não existe valor na célula.")
                elemento.value = valores_adicionar[i][j]
    except IndexError:
        print(f'Error - atualizar_dados_intervalo_planilha() | Uma ou mais células não possuem um valor para atualizar.')
    except:
        print('Error - atualizar_dados_intervalo_planilha()')
    else:
        planilha.save('PlanilhaExcel\\Arquivo.xlsx')
    finally:
        planilha.close()


def atualizar_cor_intervalo_planilha(conteudo, intervalo: str, cor: str = 'FFFFFFFF') -> None:
    """
    Atualiza a cor de fundo do intervalo informado.
    Caso seja informado apenas uma célula, apenas a cor dessa célula será atualizada.
    :param conteudo: Arquivo da planilha excel, podendo ser um path ou o arquivo em bytes.
    :param intervalo: Intervalo ou Célula a ser alterado.
    :param cor: Cor de fundo desejada no formato Hexadecimal, por padrão é branco.
    :return: None
    """
    from openpyxl.styles import PatternFill

    planilha = iniciar_planilha(conteudo)
    aba_ativa = planilha.active

    try:
        if ':' in intervalo:
            for celula in aba_ativa[intervalo]:
                celula[0].fill = PatternFill(start_color=cor, end_color=None, fill_type='solid')
        else:
            # Aplica a formatação de preenchimento à célula A1 com cor sólida
            aba_ativa[intervalo].fill = PatternFill(start_color=cor, end_color=None, fill_type='solid')
    except:
        print('Error - atualizar_cor_intervalo_planilha()')
    else:
        planilha.save('PlanilhaExcel\\Arquivo.xlsx')
    finally:
        planilha.close()
