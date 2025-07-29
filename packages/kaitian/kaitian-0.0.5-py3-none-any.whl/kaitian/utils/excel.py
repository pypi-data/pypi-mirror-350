"""Excel utils.

Developing based on `openpyxl` and `xlsxwriter`.
"""

from __future__ import annotations

import logging
from colorsys import hls_to_rgb, rgb_to_hls
from pathlib import Path
from typing import Any

import polars as pl
from typing_extensions import Self

try:
    import openpyxl
    import xlsxwriter
    import xlsxwriter.format
    import xlsxwriter.worksheet
    from openpyxl.cell.cell import Cell, MergedCell
    from openpyxl.styles.colors import Color
    from openpyxl.utils import coordinate_to_tuple, get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.xml.functions import QName, fromstring
    from xlsxwriter.utility import xl_cell_to_rowcol, xl_rowcol_to_cell
except ModuleNotFoundError as _:
    raise ModuleNotFoundError(
        "The 'excel' module requires optional dependencies 'openpyxl' and 'xlsxwriter'. "
        "You can install them using the following command:\n"
        "pip install kaitian[excel]"
    )


class ThemeParser:
    HLSMAX = 240
    XLMNS = "http://schemas.openxmlformats.org/drawingml/2006/main"

    def __init__(self, workbook: openpyxl.Workbook | Path | str) -> None:
        if isinstance(workbook, (Path, str)):
            self.workbook = openpyxl.load_workbook(Path(workbook), read_only=True)
        else:
            self.workbook = workbook

        self._parse_theme_color()

    def _parse_theme_color(self) -> None:
        _color_scheme = (
            fromstring(self.workbook.loaded_theme)
            .find(QName(self.XLMNS, "themeElements").text)
            .findall(QName(self.XLMNS, "clrScheme").text)[0]
        )

        self.colors = []

        for _c in [
            "lt1",
            "dk1",
            "lt2",
            "dk2",
            "accent1",
            "accent2",
            "accent3",
            "accent4",
            "accent5",
            "accent6",
        ]:
            accent = _color_scheme.find(QName(self.XLMNS, _c).text)
            for i in list(accent):  # walk all child nodes, rather than assuming [0]
                if "window" in i.attrib["val"]:
                    self.colors.append(i.attrib["lastClr"])
                else:
                    self.colors.append(i.attrib["val"])

    def _rgb2hls(self, rgb: str) -> tuple[int, int, int]:
        """From RGB to HLS
        Converts rgb values in range (0,1) or a hex string of the form
        '[#aa]rrggbb' to HLSMAX based HLS, alpha values are ignored.
        """
        if len(rgb) > 6:
            rgb = rgb[-6:]
        red = int(rgb[0:2], 16) / 255
        green = int(rgb[2:4], 16) / 255
        blue = int(rgb[4:6], 16) / 255

        _h, _l, _s = rgb_to_hls(red, green, blue)
        return (
            int(round(_h * self.HLSMAX)),
            int(round(_l * self.HLSMAX)),
            int(round(_s * self.HLSMAX)),
        )

    def get_theme_color(self, theme: int = 0, tint: float = 0.0) -> str:
        argb_main = self.colors[theme]
        hue, lightness, saturation = self._rgb2hls(argb_main)

        # See: http://ciintelligence.blogspot.co.uk/2012/02/converting-excel-theme-color-and-tint.html
        if tint < 0:
            lightness = int(round(lightness * (1.0 + tint)))
        else:
            lightness = int(
                round(
                    lightness * (1.0 - tint)
                    + (self.HLSMAX - self.HLSMAX * (1.0 - tint))
                )
            )

        r, g, b = hls_to_rgb(
            hue / self.HLSMAX, lightness / self.HLSMAX, saturation / self.HLSMAX
        )
        return (
            "#%02x%02x%02x"
            % (
                int(round(r * 255)),
                int(round(g * 255)),
                int(round(b * 255)),
            )
        ).upper()

    def get_indexed_color(self, index: int = 0) -> str:
        # TODO: 获取标准色
        raise NotImplementedError("get_indexed_color not implemented yet")


class FormatParser:
    def __init__(self, workbook: openpyxl.Workbook | Path | str) -> None:
        if isinstance(workbook, (Path, str)):
            self.workbook = openpyxl.load_workbook(Path(workbook))
        else:
            self.workbook = workbook

        # theme parser
        self.workbook_theme = ThemeParser(self.workbook)

        self._logger = logging.getLogger(__name__)

    def _color2rgb(self, color: Color | None, default: str = "white") -> str:
        if default.lower() == "white":
            default_rgb = "#FFFFFF"
        elif default.lower() == "black":
            default_rgb = "#000000"
        else:
            default_rgb = default

        if color is None:
            self._logger.warning(f"input color is None, use default color {default}")
            return default_rgb

        if color.type == "rgb":
            return f"#{color.rgb[2:]}"
        elif color.type == "theme":
            return self.workbook_theme.get_theme_color(color.theme, color.tint)
        else:
            self._logger.warning(f"Unknown color type: {color.type}")
            return default_rgb

    def _get_worksheet(self, sheet: str | int) -> Worksheet:
        if isinstance(sheet, str):
            worksheet = self.workbook[sheet]
        elif isinstance(sheet, int):
            worksheet = self.workbook.worksheets[sheet]
        else:
            raise ValueError("sheet must be str or int")

        return worksheet

    def _get_position(self, position: str | tuple[int, int]) -> tuple[int, int]:
        if isinstance(position, str):
            return coordinate_to_tuple(position)
        else:
            return position

    def _get_font(self, workcell: Cell | MergedCell) -> dict[str, Any]:
        font = workcell.font
        font_fmt = {
            "font_name": font.name,
            "font_size": font.size,
            "font_color": self._color2rgb(font.color, default="black"),
            "halign": workcell.alignment.horizontal,
            "valign": workcell.alignment.vertical,
            "auto_wrap": workcell.alignment.wrap_text or False,
            "background_color": "white",
        }

        if getattr(workcell.fill, "fgColor") is not None:
            font_fmt["background_color"] = self._color2rgb(
                workcell.fill.fgColor,  # type: ignore
                default="white",
            )

        return font_fmt

    def _get_shape(
        self, worksheet: Worksheet, position: tuple[int, int]
    ) -> tuple[float, float]:
        coldim = worksheet.column_dimensions[get_column_letter(position[0])]
        rowdim = worksheet.row_dimensions[position[1]]
        if getattr(coldim, "width") is not None:
            width = coldim.width
        else:
            width = coldim.style

        if getattr(rowdim, "height") is not None:
            height = rowdim.height
        else:
            height = rowdim.s

        return width, height

    def _get_border(self, workcell: Cell | MergedCell) -> dict[str, Any]:
        # cell border of (style, color)
        border = {
            direction: (
                getattr(workcell.border, direction).style,
                self._color2rgb(workcell.border.left.color),
            )
            for direction in ["top", "bottom", "left", "right"]
        }
        return border

    def _get_value_fmt(self, workcell: Cell | MergedCell) -> dict[str, Any]:
        # TODO: 获取单元格内容格式, 主要是数字和日期
        raise NotImplementedError("get_value_fmt formatter not implemented yet.")

    def _get_conditional(self, workcell: Cell | MergedCell) -> dict[str, Any]:
        # TODO: 解析条件格式
        raise NotImplementedError("get_conditional formatter not implemented yet.")

    def get_format(
        self, sheet: str | int, cell: str | tuple[int, int]
    ) -> dict[str, Any]:
        worksheet = self._get_worksheet(sheet)
        position = self._get_position(cell)

        max_row, max_col = worksheet.max_row, worksheet.max_column
        if position[0] > max_row or position[1] > max_col:
            self._logger.warning(
                f"The cell {cell} is out of range, max_row={max_row}, max_col={max_col}"
            )
            return {}

        workcell = worksheet.cell(*position)

        font_fmt = self._get_font(workcell)
        width, height = self._get_shape(worksheet, position)
        border_fmt = self._get_border(workcell)

        return dict(
            font=font_fmt, shape={"width": width, "height": height}, border=border_fmt
        )


# TODO: 将CellFormat标准化


class ExcelReader: ...


class ExcelWriter:
    def __init__(self, file_path: str | Path) -> None:
        self.workbook = xlsxwriter.Workbook(Path(file_path))
        self.formatters: dict[str, xlsxwriter.format.Format] = dict()

    def _get_worksheet_or_create(
        self, sheet_name: str
    ) -> xlsxwriter.worksheet.Worksheet:
        if sheet_name in self.workbook.sheetnames:
            worksheet = self.workbook.get_worksheet_by_name(sheet_name)
        else:
            worksheet = self.workbook.add_worksheet(sheet_name)

        assert worksheet is not None
        return worksheet

    def _get_format(self, cell_format: dict | str) -> xlsxwriter.format.Format | None:
        if isinstance(cell_format, str):
            return self.formatters.get(cell_format)

        _xlsx_fmt = {
            **(cell_format.get("font") or {}),
            "align": (cell_format.get("halign") or "left"),
            "valign": (cell_format.get("valign") or "bottom"),
            "text_wrap": (cell_format.get("auto_wrap") or False),
            "bg_color": (cell_format.get("bg_color") or "white"),
        }

        if cell_format.get("border") is not None:
            for direction in ["left", "right", "top", "bottom"]:
                b, c = cell_format["border"].get(direction)
                if b is not None:
                    _xlsx_fmt[f"{direction}"] = b
                    _xlsx_fmt[f"{direction}_color"] = c

        return self.workbook.add_format(_xlsx_fmt)

    def add_format(self, format_name: str, format_dict: dict[str, Any]) -> Self:
        fmt = self._get_format(format_dict)
        if fmt is not None:
            self.formatters[format_name] = fmt
        return self

    def fill_value(
        self,
        sheet_name: str,
        value: Any,
        position: str | tuple[int, int],
        cell_format: dict | str | None = None,
    ) -> Self:
        if isinstance(position, tuple):
            position = xl_rowcol_to_cell(*position)
        worksheet = self._get_worksheet_or_create(sheet_name)
        if cell_format is not None:
            fmt = self._get_format(cell_format)
        else:
            fmt = None
        worksheet.write(position, value, fmt)
        return self

    def fill_table(
        self,
        sheet_name: str,
        data: pl.DataFrame,
        header: bool = True,
        index: bool = False,
        offset: str | tuple[int, int] = "A1",
        column_format: dict | None = None,
    ) -> Self:
        if index:
            data = data.with_row_index(offset=1)

        worksheet = self._get_worksheet_or_create(sheet_name)

        if isinstance(offset, str):
            offset = xl_cell_to_rowcol(offset)

        # TODO: 添加格式
        _ = worksheet.write_row(*offset, data.columns)
        for col_idx, series in enumerate(data):
            worksheet.write_column(
                row=offset[0] + 1, col=offset[1] + col_idx, data=series.to_list()
            )
        return self

    def merge_cell(
        self,
        sheet_name: str,
        offset: str | tuple[int, int],
        nrows: int,
        ncols: int,
        value: Any = None,
    ) -> Self:
        if isinstance(offset, str):
            offset = xl_cell_to_rowcol(offset)

        worksheet = self._get_worksheet_or_create(sheet_name)
        _ = worksheet.merge_range(
            *offset,
            last_row=offset[0] + nrows - 1,
            last_col=offset[1] + ncols - 1,
            data=value,
        )
        return self

    def done(self) -> None:
        self.workbook.close()
