from typing import Any, TypeAlias, TextIO, BinaryIO
# from typing_extensions import TypeIs

import openpyxl as op

from ka_uts_log.log import LogEq
from ka_uts_obj.io import Io
from ka_uts_xls.op.wbop import WbOp
from ka_uts_xls.op.wsop import WsOp

from pathlib import Path

TyWbOp: TypeAlias = op.workbook.workbook.Workbook
TyWsOp: TypeAlias = op.worksheet.worksheet.Worksheet

TyArr = list[Any]
TyDic = dict[Any, Any]
TyAoA = list[TyArr]
TyAoD = list[TyDic]
TyDoAoA = dict[Any, TyAoA]
TyDoAoD = dict[Any, TyAoD]
TyDoWsOp = dict[Any, TyWsOp]
TyAoD_DoAoD = TyAoD | TyDoAoD
TyOpFileSrc = str | bytes | Path | TextIO | BinaryIO
# TyOpFileSrc = str | bytes | TyXls | Path | TextIO | BinaryIO
TyPath = str
TyPathnm = str

TySheet = int | str
TySheets = int | str | list[int | str]
TySheetname = str
TySheetnames = list[TySheetname]

TnArr = None | TyArr
TnAoA = None | TyAoA
TnAoD = None | TyAoD
TnAoD_DoAoD = None | TyAoD_DoAoD
TnDic = None | TyDic
TnDoAoA = None | TyDoAoA
TnDoAoD = None | TyDoAoD
TnDoWsOp = None | TyDoWsOp
TnSheet = None | TySheet
TnSheets = None | TySheets
TnSheetname = None | TySheetname
TnSheetnames = None | TySheetnames
TnWsOp = None | TyWsOp
# TnDf_DoDf = TnPdDf_DoPdDf | TnPlDf_DoPlDf


class IoiPathWbOp:

    @staticmethod
    def load(io: TyOpFileSrc, **kwargs) -> TyWbOp:
        if io == '':
            raise Exception('io is empty String')
        if io is None:
            raise Exception('io is None')
        try:
            wb: TyWbOp = op.load_workbook(io, **kwargs)
        except Exception as e:
            msg = f"openpyxl.load_workbook for io = {io!r} throw exception {e}"
            raise Exception(msg)
        return wb

    @classmethod
    def read_wb_to_aod(cls, io: TyOpFileSrc, sheet: TnSheet, **kwargs) -> TyAoD:
        Io.verify(io)
        _wb: TyWbOp = cls.load(io, **kwargs)
        return WbOp.to_aod(_wb, sheet)

    @staticmethod
    def read_wb_to_doaod(
            io: TyOpFileSrc, sheet: TnSheets, **kwargs) -> TyDoAoD:
        Io.verify(io)
        _wb: TyWbOp = IoiPathWbOp.load(io, **kwargs)
        return WbOp.to_doaod(_wb, sheet)

    @staticmethod
    def read_wb_to_aod_or_doaod(
            io: TyOpFileSrc, sheet: TnSheets, **kwargs) -> TnAoD_DoAoD:
        Io.verify(io)
        _wb: TyWbOp = IoiPathWbOp.load(io, **kwargs)
        return WbOp.to_aod_or_doaod(_wb, sheet)

    @staticmethod
    def read_wb_to_aoa(io: TyOpFileSrc, **kwargs) -> tuple[TyAoA, TyAoA]:
        Io.verify(io)
        wb: TyWbOp = IoiPathWbOp.load(io)
        heads_sheet_name = kwargs.get('headers_sheet_name')
        ws_names: TySheetnames = WbOp.sh_sheetnames(wb, **kwargs)
        aoa = []
        if heads_sheet_name is not None:
            ws = wb[heads_sheet_name]
            heads = WsOp.sh_headers(ws, **kwargs)
        else:
            heads = []
        for ws_name in ws_names:
            LogEq.debug("ws_name", ws_name)
            ws = wb[ws_name]
            aoa_ws = WsOp.sh_aoa(ws, sheet_name=ws_name, **kwargs)
            aoa.extend(aoa_ws)
            LogEq.debug("aoa_ws", aoa_ws)
        return heads, aoa

    @classmethod
    def read_wb_to_aoa_by_prefix(cls, **kwargs) -> TyAoA:
        # ex_read_workbook_2_aoa(cls, **kwargs):
        # def ex_read_aoa(cls, **kwargs):
        prefix = kwargs.get('prefix')
        if prefix is not None:
            prefix = f"_{prefix}"
        in_io: TyOpFileSrc = kwargs.get(f'in_path{prefix}', '')
        row_start = kwargs.get(f'row_start{prefix}')
        cols_count = kwargs.get(f'cols_count{prefix}')
        sw_add_sheet_name = kwargs.get(f'sw_add_sheet_name{prefix}')
        sheet_names = kwargs.get(f'sheet_names{prefix}')
        headers_sheet_name = kwargs.get(f'headers_sheet_name{prefix}')
        headers_start = kwargs.get(f'headers_start{prefix}')
        Io.verify(in_io)
        heads, aoa = cls.read_wb_to_aoa(
                in_io,
                row_start=row_start,
                cols_count=cols_count,
                sw_add_sheet_name=sw_add_sheet_name,
                sheet_names=sheet_names,
                headers_sheet_name=headers_sheet_name,
                headers_start=headers_start)
        return aoa


class IoiPathWsOp:

    @staticmethod
    def read_ws_to_dic(
            io: TyOpFileSrc, sheet: TySheet) -> TnDic:
        _wb: TyWbOp = IoiPathWbOp.load(io)
        _ws: TnWsOp = WbOp.sh_sheet(_wb, sheet)
        return WsOp.to_dic(_ws)

    @staticmethod
    def read_ws_to_aod(
            io: TyOpFileSrc, sheet: TySheet) -> TnAoD:
        _wb: TyWbOp = IoiPathWbOp.load(io)
        _ws: TnWsOp = WbOp.sh_sheet(_wb, sheet)
        return WsOp.to_aod(_ws)

    @staticmethod
    def read_ws_filter_rows(io: TyOpFileSrc, sheet: TySheet) -> TnArr:
        Io.verify(io)
        _wb: TyWbOp = IoiPathWbOp.load(io)
        _ws: TnWsOp = WbOp.sh_sheet(_wb, sheet)
        return WsOp.filter_rows(_ws)

    @staticmethod
    def read_ws_to_aoa(
            io: TyOpFileSrc, sheet: TnSheets = None) -> tuple[TnAoA, TnSheetnames]:
        Io.verify(io)
        _wb: TyWbOp = IoiPathWbOp.load(io)
        aoa: TyAoA = []
        if not sheet:
            return aoa, None
        _sheetnames: TnSheetnames = WbOp.sh_sheetnames(_wb, sheet)
        if not _sheetnames:
            return aoa, _sheetnames
        for _sheetname in _sheetnames:
            _ws: TnWsOp = WbOp.sh_worksheet(_wb, _sheetname)
            if _ws is not None:
                values: TyArr = WsOp.to_row_values(_ws)
                aoa.append(values)
        return aoa, _sheetnames

    @staticmethod
    def read_sheetnames(io: TyOpFileSrc) -> TyArr:
        Io.verify(io)
        wb: TyWbOp = IoiPathWbOp.load(io)
        sheetnames: TySheetnames = wb.sheetnames
        return sheetnames

    @staticmethod
    def read_ws_to_doaoa(
            io: TyOpFileSrc, sheet: TnSheets = None) -> tuple[TnDoAoA, TnSheetnames]:
        Io.verify(io)
        _wb: TyWbOp = IoiPathWbOp.load(io)
        doaoa: TyDoAoA = {}
        if _wb is None:
            return doaoa, None
        sheetnames: TnSheetnames = WbOp.sh_sheetnames(_wb, sheet)
        if not sheetnames:
            return doaoa, sheetnames
        for _sheetname in sheetnames:
            _ws: TnWsOp = WbOp.sh_worksheet(_wb, _sheetname)
            if _ws is not None:
                values: TyArr = WsOp.to_row_values(_ws)
                doaoa[sheet] = values
        return doaoa, sheetnames

    @staticmethod
    def read_ws_to_dowsop(
            io: TyOpFileSrc, sheet: TnSheets = None) -> tuple[TnDoWsOp, TnSheetnames]:
        Io.verify(io)
        _wb: TyWbOp = IoiPathWbOp.load(io)
        dows: TyDoWsOp = {}
        if _wb is None:
            return dows, None
        sheetnames: TnSheetnames = WbOp.sh_sheetnames(_wb, sheet)
        if not sheetnames:
            return dows, sheetnames
        for _sheetname in sheetnames:
            _ws: TnWsOp = WbOp.sh_worksheet(_wb, _sheetname)
            if _ws is not None:
                dows[_sheetname] = _ws
        return dows, sheetnames
