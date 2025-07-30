from typing import Any, TypeAlias

import openpyxl as op
import pandas as pd

from ka_uts_aod.aod import AoD
from ka_uts_dic.dic import Dic
from ka_uts_xls.op.ioc import IocWbOp
from ka_uts_xls.op.wsop import WsOp

TyOpCe: TypeAlias = op.cell.cell.Cell
TyWbOp: TypeAlias = op.workbook.workbook.Workbook
TyWsOp: TypeAlias = op.worksheet.worksheet.Worksheet
TyCsOp: TypeAlias = op.chartsheet.chartsheet.Chartsheet
TyPdDf: TypeAlias = pd.DataFrame

# TyWsCsOp = Worksheet | WriteOnlyWorksheet | ReadOnlyWorksheet | Chartsheet

TyArr = list[Any]
TyAoA = list[TyArr]
TyAoAoA = list[TyAoA]
TyDic = dict[Any, Any]
TyAoD = list[TyDic]
TyAoS = list[str]
TyAoWsOp = list[TyWsOp]
TyDoD = dict[Any, TyDic]
TyDoAoA = dict[Any, TyAoA]
TyDoAoD = dict[Any, TyAoD]
TyDoWsOp = dict[Any, TyWsOp]
TyDoPdDf = dict[Any, TyPdDf]
TyAoD_DoAoD = TyAoD | TyDoAoD
TySheet = int | str
TyOpSheets = TySheet | list[int | str]
TySheetname = str
TySheetnames = list[TySheetname]
TyStrArr = str | TyArr
TyTupleOpCe = tuple[TyOpCe, ...]

TnArr = None | TyArr
TnAoA = None | TyAoA
TnAoD = None | TyAoD
TnDic = None | TyDic
TnDoAoA = None | TyDoAoA
TnAoD_DoAoD = None | TyAoD_DoAoD
TnAoWsOp = None | TyAoWsOp
TnDoWsOp = None | TyDoWsOp
TnOpSheets = None | TyOpSheets
TnSheet = None | TySheet
TnSheetname = None | TySheetname
TnWbOp = None | TyWbOp
TnWsOp = None | TyWsOp
TnCsOp = None | TyCsOp


class WbOp:

    # @staticmethod
    # def get(**kwargs) -> TyWbOp:
    #     wb: TyWbOp = Openpyxl_.get_workbook(write_only=True)
    #     return wb

    @staticmethod
    def create_wb_with_doaoa(doaoa: TnDoAoA) -> TyWbOp:
        wb: TyWbOp = IocWbOp.get(write_only=True)
        if not doaoa:
            ws: None | TyWsOp | TyCsOp = wb.active
            if ws is None:
                return wb
            wb.remove(ws)
            return wb
        for ws_id, aoa in doaoa.items():
            _ws: None | TyWsOp = wb.create_sheet()
            if _ws is None:
                continue
            _ws.title = ws_id
            WsOp.append_rows(_ws, aoa)
        return wb

    @staticmethod
    def create_wb_from_doaod(doaod: TyDoAoD) -> TyWbOp:
        wb: TyWbOp = IocWbOp.get(write_only=True)
        if not doaod:
            # wb.remove(ws)
            return wb
        for ws_id, aod in doaod.items():
            a_header = [list(aod[0].keys())]
            a_data = [list(d.values()) for d in aod]
            a_row = a_header + a_data
            ws: TyWsOp = wb.create_sheet()
            ws.title = ws_id
            WsOp.append_rows(ws, a_row)
        return wb

    @staticmethod
    def iter_sheet_names(wb: TyWbOp, **kwargs):
        cols_count = kwargs.get('cols_count', 0)
        sheet_names: TyArr = kwargs.get('sheet_names', [])
        for sheet_name in sheet_names:
            sheet = wb[sheet_name]
            if sheet.max_column == cols_count:
                yield sheet_name

    @staticmethod
    def iter_sheet(wb: TyWbOp, max_sheets):
        for _ii in range(0, max_sheets):
            yield wb.create_sheet()

    @staticmethod
    def sh_sheetname_in_arr(sheet: TnSheet, sheetnames: TySheetnames) -> TnSheetname:
        if not sheet:
            return None
        if isinstance(sheet, int):
            if sheet < len(sheetnames):
                return sheetnames[sheet]
            return None
        if isinstance(sheet, str):
            if sheet in sheetnames:
                return sheet
            return None
        return None

    @classmethod
    def sh_sheetname(cls, wb: TnWbOp, sheet: TnSheet) -> TnSheetname:
        if wb is None:
            return None
        return cls.sh_sheetname_in_arr(sheet, wb.sheetnames)

    @classmethod
    def sh_sheetnames(cls, wb: TnWbOp, sheet: TnOpSheets) -> TySheetnames:
        if wb is None:
            return []
        if not sheet:
            sheetnames: TySheetnames = wb.sheetnames
            return sheetnames
        if isinstance(sheet, (int, str)):
            sheetname: TnSheetname = cls.sh_sheetname_in_arr(sheet, wb.sheetnames)
            if not sheetname:
                return []
            return [sheetname]
        if isinstance(sheet, (list, tuple)):
            sheetnames_new = []
            for _sheet in sheet:
                _sheetname = cls.sh_sheetname_in_arr(_sheet, wb.sheetnames)
                if _sheetname:
                    sheetnames_new.append(_sheetname)
            return sheetnames_new
        return []

    @classmethod
    def sh_sheet_by_sheetname(
            cls, wb: TnWbOp, sheetname: TnSheetname) -> TnWsOp:
        if wb is None:
            return None
        if not sheetname:
            return None
        return wb[sheetname]

    @classmethod
    def sh_sheet(cls, wb: TnWbOp, sheet: TySheet) -> TnWsOp:
        return cls.sh_sheet_by_sheetname(wb, cls.sh_sheetname(wb, sheet))

    @classmethod
    def sh_sheet_by_type(
            cls, wb: TnWbOp, sheet_name: TnSheetname, sheet_type: str
    ) -> TnWsOp | TyCsOp:
        return WsOp.sh_by_type(cls.sh_sheet_by_sheetname(wb, sheet_name), sheet_type)

    @classmethod
    def sh_chartsheet_by_sheetname(cls, wb: TnWbOp, sheet_name: TnSheetname) -> TnCsOp:
        return WsOp.sh_chartsheet(cls.sh_sheet_by_sheetname(wb, sheet_name))

    @classmethod
    def sh_worksheet_by_sheetname(cls, wb: TnWbOp, sheet_name: TnSheetname) -> TnWsOp:
        return WsOp.sh_worksheet(cls.sh_sheet_by_sheetname(wb, sheet_name))

    @classmethod
    def sh_chartsheet(cls, wb: TnWbOp, sheet: TnSheet) -> TnCsOp:
        return cls.sh_chartsheet_by_sheetname(wb, cls.sh_sheetname(wb, sheet))

    @classmethod
    def sh_worksheet(cls, wb: TnWbOp, sheet: TnSheet) -> TnWsOp:
        return cls.sh_worksheet_by_sheetname(wb, cls.sh_sheetname(wb, sheet))

    @classmethod
    def to_aod(cls, wb: TnWbOp, sheet: TnSheet) -> TyAoD:
        if wb is None:
            return []
        _ws: TnWsOp = cls.sh_worksheet(wb, sheet)
        return WsOp.to_aod(_ws)

    @classmethod
    def to_doaod(cls, wb: TnWbOp, sheet: TnOpSheets) -> TyDoAoD:
        if wb is None:
            return {}
        doaod: TyDoAoD = {}
        if wb is None:
            return doaod
        _sheetnames: TySheetnames = cls.sh_sheetnames(wb, sheet)
        if not _sheetnames:
            return doaod
        for _sheetname in _sheetnames:
            _ws: TnWsOp = cls.sh_worksheet_by_sheetname(wb, _sheetname)
            Dic.set_kv_not_none(doaod, _sheetname, WsOp.to_aod(_ws))
        return doaod

    @classmethod
    def to_aod_or_doaod(
            cls, wb: TyWbOp, sheet: TnOpSheets) -> TyAoD_DoAoD:
        doaod: TyDoAoD = {}
        _sheetnames: TySheetnames = cls.sh_sheetnames(wb, sheet)
        if not _sheetnames:
            return doaod
        if len(_sheetnames) == 1:
            _sheetname = _sheetnames[0]
            _ws: TnWsOp = WbOp.sh_worksheet_by_sheetname(wb, _sheetname)
            return WsOp.to_aod(_ws)
        for _sheetname in _sheetnames:
            _ws = WbOp.sh_worksheet_by_sheetname(wb, _sheetname)
            Dic.set_kv_not_none(doaod, _sheetname, WsOp.to_aod(_ws))
        return doaod

    @classmethod
    def createupdate_wb_with_doaoa(cls, wb: TnWbOp, doaoa: TnDoAoA) -> None:
        if not doaoa:
            return
        if wb is None:
            cls.create_wb_with_doaoa(doaoa)
        else:
            cls.update_wb_with_doaoa(wb, doaoa)

    @classmethod
    def update_wb_with_aoa(cls, wb: TnWbOp, aoa: TnAoA, sheet_name: str) -> None:
        if wb is None:
            return
        if not aoa:
            return
        _sheet_name: TnSheetname = cls.sh_sheetname(wb, sheet_name)
        _ws: TnWsOp = cls.sh_worksheet_by_sheetname(wb, _sheet_name)
        WsOp.append_rows(_ws, aoa)

    @classmethod
    def update_wb_with_aod(cls, wb: TnWbOp, aod: TnAoD, sheet_name: str) -> None:
        if wb is None:
            return
        _aoa: TnAoA = AoD.to_aoa(aod, sw_keys=False)
        cls.update_wb_with_aoa(wb, _aoa, sheet_name)

    @classmethod
    def update_wb_with_doaoa(cls, wb: TnWbOp, doaoa: TnDoAoA) -> None:
        if wb is None:
            return
        if not doaoa:
            return
        a_ws_id: TyArr = Dic.sh_keys(doaoa, wb.sheetnames)
        for ws_id in a_ws_id:
            aoa: TyAoA = doaoa[ws_id]
            ws: TnWsOp = cls.sh_worksheet(wb, ws_id)
            WsOp.append_rows(ws, aoa)

    @staticmethod
    def update_wb_with_dodf(wb: TnWbOp, dodf: TyDoPdDf, **kwargs) -> TnWbOp:
        if wb is None:
            return None
        _d_update: TyDic = kwargs.get('d_update', {})
        _d_head: TyDic = _d_update.get('d_head', {})
        _a_key: TyArr = Dic.show_sorted_keys(dodf)
        for _key in _a_key:
            _df = dodf[_key]
            _ws_tpl: TyWsOp = wb['TMPL']
            _ws_new: TyWsOp = wb.copy_worksheet(_ws_tpl)
            _ws_new.title = _key
            _d_head['title']['value'] = _key
            WsOp.update_ws_cell_from_df_with_d_body(_ws_new, _df, _d_update)
            WsOp.update_ws_cell_with_d_head(_ws_new, _d_head)
        return wb
