import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, Side, Border
from typing import List, Dict, Union


def format_excel(df: pd.DataFrame, file_path: str, precision: int = 2) -> None:
    """Apply Excel formatting to a DataFrame and save to file."""
    wb = Workbook()
    ws = wb.active

    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    print("\nFormatting Excel: adjusting column width and precision...")

    # Header formatting
    for col_idx in range(1, df.shape[1] + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(name="Arial", size=10, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(border_style="thin", color='FF000000'))

    # Content formatting
    for row_idx in range(2, df.shape[0] + 2):
        for col_idx in range(1, df.shape[1] + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")

            if isinstance(cell.value, int):
                cell.number_format = '0'
            elif isinstance(cell.value, float):
                cell.number_format = f'0.{"0" * precision}'

    # Column width setting
    for col_idx in range(1, ws.max_column + 1):
        width = 25 if col_idx == 1 else 15
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(file_path)


def init_table(columns: List[str]) -> Dict[str, List]:
    """Initialize an empty table (dict) from column names."""
    return {col: [] for col in columns}


def write_table(data: Dict[str, List],
                return_df: bool = False,
                save_excel: bool = True,
                excel_path: Union[str, None] = None,
                formatting: bool = True,
                precision: int = 2) -> Union[None, pd.DataFrame]:
    """Write table data to Excel and optionally return DataFrame."""
    if not isinstance(data, dict):
        raise TypeError(f"Expected dict, got {type(data)}")

    df = pd.DataFrame(data)

    if save_excel and excel_path:
        if formatting:
            format_excel(df, excel_path, precision)
        else:
            df.to_excel(excel_path, index=False)
        print(f"\nSaved Excel to: {excel_path}\n")

    return df if return_df else None


def build_dataframe(rows: List[List], columns: List[str],
                    return_df: bool = False,
                    save_excel: bool = True,
                    excel_path: Union[str, None] = None,
                    formatting: bool = True,
                    precision: int = 2) -> Union[None, pd.DataFrame]:
    """Build and optionally save a DataFrame from row data and column names."""
    if not (isinstance(rows, list) and isinstance(columns, list)):
        raise TypeError(f"Both inputs must be lists. Got {type(rows)} and {type(columns)}")

    if not rows or len(rows[0]) != len(columns):
        raise ValueError("Mismatch between number of columns and row values")

    table = init_table(columns)
    for row in rows:
        for col, val in zip(columns, row):
            table[col].append(val)

    return write_table(table, return_df, save_excel, excel_path, formatting, precision)


if __name__ == '__main__':
    
    # Example usage
    sample_rows = [[1, 2, 3], [4, 5, 6], [7, 8, 9]]
    sample_columns = ['a', 'b', 'c']
    df_result = build_dataframe(sample_rows, sample_columns, excel_path='test.xlsx',
                                save_excel=True, return_df=True)
    print(df_result)
