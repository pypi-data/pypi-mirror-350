#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Fp4NT - HTML Table to Markdown and Excel to DataFrame Converter

提供以下功能：
    1. 将HTML表格转换为Markdown格式（支持合并单元格）
    2. 将Excel文件中的工作表读取为Pandas DataFrame（支持合并单元格）

作者: lrs33 
邮箱: aslongrushan@gmail.com
创建日期: 2025-04-24
最后修改日期: 2025-05-26
版本: 1.0.0
Python版本要求: >=3.7
依赖库:
    - beautifulsoup4>=4.9.0
    - html2text>=2020.1.16
    - pandas>=1.0.0
    - openpyxl>=3.0.0
"""

import logging
from datetime import datetime
from bs4 import BeautifulSoup
from html2text import html2text
import pandas as pd
from openpyxl import load_workbook


class Fp4NT:
    """
    主类 Fp4NT（Full Processing for Nested Tables），用于统一处理HTML表格和Excel表格。
    
    支持功能：
        - HTML 表格转 Markdown
        - Excel 表格转 Pandas DataFrame（带合并单元格处理）
    
    参数：
        row_fill_merged (bool): 是否自动按行填充合并单元格，默认为 True
        col_fill_merged (bool): 是否自动按列填充合并单元格，默认为 False
        fill_mark (str): 合并单元格的填充标记，默认为 '无'
        verbose (bool): 是否输出日志信息，默认为 True
    """

    def __init__(self, row_fill_merged=True, col_fill_merged=False, fill_mark='无', verbose=True):
        self.row_fill_merged = row_fill_merged
        self.col_fill_merged = col_fill_merged
        self.fill_mark = fill_mark
        self.verbose = verbose
        self._setup_logging()

    def _setup_logging(self):
        """设置日志系统，便于调试和跟踪程序运行状态"""
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
        )
        self.logger = logging.getLogger(__name__)

    def convert_html_to_markdown(self, html_text):
        """
        将包含HTML表格的文本转换为Markdown格式
        
        参数：
            html_text (str): 包含HTML表格的字符串
            
        返回：
            str: 转换后的Markdown文本
        """
        start_time = datetime.now()
        if self.verbose:
            self.logger.info(f"开始处理HTML文本，长度: {len(html_text)} 字符")

        try:
            soup = BeautifulSoup(html_text, 'html.parser')
            tables = soup.find_all('table')

            if self.verbose:
                self.logger.info(f"找到 {len(tables)} 个表格")

            # 倒序处理表格以避免替换时干扰后续内容
            for i, table in enumerate(reversed(tables), 1):
                if self.verbose:
                    self.logger.debug(f"正在处理第 {i}/{len(tables)} 个表格")

                markdown_table = self._convert_single_table(table)
                replacement = soup.new_tag("pre")
                replacement.string = markdown_table
                table.replace_with(replacement)

            reparse_html_str = str(soup.prettify())
            result = html2text(reparse_html_str)

            if self.verbose:
                elapsed = (datetime.now() - start_time).total_seconds()
                self.logger.info(f"处理完成，耗时: {elapsed:.2f} 秒")

            return result

        except Exception as e:
            self.logger.error(f"处理HTML文本时出错: {str(e)}")
            raise

    def _convert_single_table(self, table):
        """
        将单个HTML表格转换为Markdown格式
        
        参数：
            table (bs4.element.Tag): BeautifulSoup表格对象
            
        返回：
            str: Markdown格式的表格字符串
        """
        if not table:
            if self.verbose:
                self.logger.warning("传入的表格对象为空")
            return ""

        start_time = datetime.now()
        if self.verbose:
            self.logger.debug("开始处理单个表格")

        try:
            rows = []
            for tr in table.find_all('tr'):
                row = []
                exist_text = False
                rowspan_counts = []
                for cell in tr.find_all(['th', 'td']):
                    # 获取单元格文本并清理
                    cell_text = ' '.join(cell.get_text(' ', strip=True).split())
                    if cell_text.replace(" ", ""):
                        exist_text = True

                    # 处理rowspan和colspan属性
                    rowspan = int(cell.get('rowspan', 1))
                    colspan = int(cell.get('colspan', 1))

                    # 存储单元格信息
                    row.append({
                        'text': cell_text,
                        'rowspan': rowspan,
                        'colspan': colspan,
                        'is_empty': not bool(cell_text) and not cell.find(True)
                    })
                    rowspan_counts.append(rowspan)

                if exist_text:
                    # 如果该行所有单元格跨行数一致，则更新为1
                    if len(set(rowspan_counts)) == 1:
                        for r in row:
                            r['rowspan'] = 1
                    rows.append(row)

            if not rows:
                if self.verbose:
                    self.logger.warning("表格中没有找到有效行")
                return ""

            # 计算最大列数
            max_cols = max(sum(cell['colspan'] for cell in row) for row in rows)

            if self.verbose:
                self.logger.debug(f"表格结构 - 行数: {len(rows)}, 最大列数: {max_cols}")

            # 创建填充网格
            grid = [[{'type': 'unfilled', 'value': ''} for _ in range(max_cols)] for _ in range(len(rows))]

            # 填充网格数据
            for i, row in enumerate(rows):
                col_pos = 0
                for cell in row:
                    while col_pos < max_cols and grid[i][col_pos]['type'] != 'unfilled':
                        col_pos += 1
                    if col_pos >= max_cols:
                        break

                    # 标记主单元格
                    if cell['is_empty']:
                        grid[i][col_pos] = {'type': 'original_empty', 'value': self.fill_mark}
                    else:
                        grid[i][col_pos] = {'type': 'original', 'value': cell['text']}

                    # 标记合并的单元格位置
                    for r in range(i, i + cell['rowspan']):
                        for c in range(col_pos, col_pos + cell['colspan']):
                            if r == i and c == col_pos:
                                continue
                            if r < len(grid) and c < max_cols:
                                should_fill = (
                                    (r != i and c == col_pos and self.row_fill_merged) or
                                    (r == i and c != col_pos and self.col_fill_merged) or
                                    (r != i and c != col_pos and self.row_fill_merged and self.col_fill_merged)
                                )
                                if should_fill:
                                    grid[r][c] = {
                                        'type': 'filled',
                                        'value': cell['text'] if cell['text'] else self.fill_mark
                                    }

                    col_pos += cell['colspan']

            # 生成Markdown表格
            markdown_lines = []

            # 表头行
            if len(grid) > 0:
                headers = [cell['value'] for cell in grid[0]]
                markdown_lines.append("| " + " | ".join(headers) + " |")

                # 分隔线
                markdown_lines.append("| " + " | ".join(["---"] * len(headers)) + " |")

                # 数据行
                for row in grid[1:]:
                    row_content = [cell['value'] if cell['value'] else self.fill_mark for cell in row]
                    md_content = "| " + " | ".join(row_content) + " |"
                    if md_content not in markdown_lines:
                        markdown_lines.append(md_content)

            if self.verbose:
                elapsed = (datetime.now() - start_time).total_seconds()
                self.logger.debug(f"表格处理完成，耗时: {elapsed:.4f} 秒")

            return "\n".join(markdown_lines)

        except Exception as e:
            self.logger.error(f"处理表格时出错: {str(e)}")
            raise


    def build_excel_frame(self, file_path, sheet_name=0):
        """
        将Excel文件中的指定工作表读取为DataFrame，并处理合并单元格
        
        参数：
            file_path (str): Excel文件路径
            sheet_name (str or int): 工作表名称或索引（默认为0）
            
        返回：
            pd.DataFrame: 包含处理后数据的DataFrame
        """
        wb = load_workbook(filename=file_path)
        ws = wb.worksheets[sheet_name] if isinstance(sheet_name, int) else wb[sheet_name]

        max_row = ws.max_row
        max_col = ws.max_column

        full_data = [[None for _ in range(max_col)] for _ in range(max_row)]

        merged_cells_set = set()
        merged_map = {}

        # 遍历合并区域，构建映射关系
        for merged_range in ws.merged_cells.ranges:
            min_col, min_row, max_col_merge, max_row_merge = merged_range.bounds
            min_row -= 1
            min_col -= 1
            max_row_merge -= 1
            max_col_merge -= 1
            top_left_value = ws.cell(row=min_row + 1, column=min_col + 1).value

            for r in range(min_row, max_row_merge + 1):
                for c in range(min_col, max_col_merge + 1):
                    merged_cells_set.add((r, c))
                    merged_map[(r, c)] = (min_row, min_col, top_left_value)

        # 填充数据
        for row in ws.iter_rows():
            for cell in row:
                r = cell.row - 1
                c = cell.column - 1

                if (r, c) in merged_cells_set:
                    min_row, min_col, value = merged_map[(r, c)]
                    if (r == min_row or self.row_fill_merged) and (c == min_col or self.col_fill_merged):
                        full_data[r][c] = value
                    else:
                        full_data[r][c] = self.fill_mark
                else:
                    full_data[r][c] = cell.value

        # 重构DataFrame
        header_row = next(ws.iter_rows())  # 第一行作为列名
        columns = [cell.value for cell in header_row]
        df = pd.DataFrame(full_data[1:], columns=columns)

        # 删除全为空的列
        df.dropna(axis=1, how='all', inplace=True)

        return df


if __name__ == "__main__":
    fp4nt = Fp4NT(row_fill_merged=True, col_fill_merged=False)

    # 示例1：HTML表格转Markdown
    html = "<table><tr><th>Header1</th><th>Header2</th></tr><tr><td>Data1</td><td>Data2</td></tr></table>"
    markdown_output = fp4nt.convert_html_to_markdown(html)
    print("HTML转Markdown结果：")
    print(markdown_output)

    # 示例2：Excel转DataFrame
    data_path = '/data/yfzx/lrs/fp4nt/test.xlsx'
    df = fp4nt.build_excel_frame(data_path)
    print("\nExcel转DataFrame结果：")
    print(df.to_markdown(index=False))